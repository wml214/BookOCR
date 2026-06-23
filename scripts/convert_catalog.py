"""将原始馆藏 Excel 转换为适合 OCR 纠错检索的精简 CSV。

原始馆藏表按“册”保存，同一种书可能出现多行。本脚本会：

1. 只读取 OCR 匹配真正需要的字段，降低后续加载内存；
2. 对题名生成规范化文本，便于进行模糊匹配；
3. 按书目元数据记录号合并馆藏复本，并统计每种书的册数；
4. 使用 UTF-8 BOM 写出 CSV，确保 Excel、WPS 和 Python 都能正确显示中文。

脚本只读取原始 XLSX，不会修改或覆盖原始馆藏数据。
"""

from __future__ import annotations

import argparse
import csv
import re
import unicodedata
import warnings
from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


# 以当前脚本的位置推导项目根目录，而不是依赖运行命令时的工作目录。
# 这样无论在 PyCharm 中直接运行，还是从项目根目录、scripts 目录启动，
# 默认输入和输出路径都会稳定指向同一位置。
PROJECT_ROOT = Path(__file__).resolve().parents[1]


# 原始表头位于第 5 行。这里使用中文列名而不是固定列号，
# 这样教师若调整了列顺序，只要列名不变，脚本仍可以工作。
REQUIRED_COLUMNS = (
    "元数据记录号",
    "文献类型",
    "索书号",
    "条码号",
    "题名",
    "责任者",
    "出版社",
    "出版年",
    "标准号",
    "馆藏地",
    "书刊状态",
)

# 输出列顺序同时也是 CSV 的稳定接口。后续馆藏匹配模块应按这些列名读取，
# 不要依赖某一列在 CSV 中的数字位置。
OUTPUT_COLUMNS = (
    "metadata_id",
    "title",
    "title_normalized",
    "author",
    "call_number",
    "publisher",
    "publication_year",
    "standard_number",
    "document_type",
    "holding_location",
    "available_copy_count",
    "total_copy_count",
)


@dataclass
class CatalogRecord:
    """一条经过合并的馆藏书目记录。"""

    metadata_id: str
    title: str
    title_normalized: str
    author: str
    call_number: str
    publisher: str
    publication_year: str
    standard_number: str
    document_type: str
    holding_location: str
    available_copy_count: int = 0
    total_copy_count: int = 0

    def to_row(self) -> dict[str, str | int]:
        """转换为可直接写入 ``csv.DictWriter`` 的字典。"""

        return {
            "metadata_id": self.metadata_id,
            "title": self.title,
            "title_normalized": self.title_normalized,
            "author": self.author,
            "call_number": self.call_number,
            "publisher": self.publisher,
            "publication_year": self.publication_year,
            "standard_number": self.standard_number,
            "document_type": self.document_type,
            "holding_location": self.holding_location,
            "available_copy_count": self.available_copy_count,
            "total_copy_count": self.total_copy_count,
        }


def cell_text(value: Any) -> str:
    """把 Excel 单元格值安全转换成去除首尾空白的字符串。

    Excel 中的编号可能被解析为浮点数，例如 ``281046.0``。对于这类整数型
    浮点数，应去掉无意义的 ``.0``，避免后续把同一编号视为不同字符串。
    """

    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_title(title: str) -> str:
    """生成用于 OCR 模糊匹配的书名规范化文本。

    处理规则：

    - 使用 Unicode NFKC 合并全角/半角字符；
    - 转成小写，避免英文大小写差异；
    - 只保留中文、英文字母和数字；
    - 删除空格、标点及装饰符号。

    这里不直接删除“第X版”“上册”等信息，因为它们对区分馆藏版本很重要。
    """

    normalized = unicodedata.normalize("NFKC", title).lower()
    return re.sub(r"[^0-9a-z\u4e00-\u9fff]+", "", normalized)


def choose_first(current: str, candidate: str) -> str:
    """优先保留已有非空值，已有值为空时才使用候选值。"""

    return current or candidate


def read_catalog_rows(
    source: Path,
    sheet_name: str,
    header_row: int,
) -> Iterable[dict[str, str]]:
    """以只读流式模式读取馆藏 Excel，并逐行返回需要的字段。

    原始工作簿的尺寸元数据不规范，OpenPyXL 可能误判为只有 A1 一个单元格。
    调用 ``reset_dimensions`` 后，读取器会根据实际 XML 内容重新遍历全部行。
    """

    # 原文件缺少默认样式，OpenPyXL 会产生不影响数据读取的警告。
    # 仅屏蔽这一已知警告，其他异常仍正常抛出。
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Workbook contains no default style",
            category=UserWarning,
        )
        workbook = load_workbook(source, read_only=True, data_only=True)

    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"工作表 {sheet_name!r} 不存在，可用工作表：{workbook.sheetnames}"
            )

        worksheet = workbook[sheet_name]
        worksheet.reset_dimensions()
        rows = worksheet.iter_rows(values_only=True)

        # 跳过说明行，直到指定的表头行。
        header: tuple[Any, ...] | None = None
        for row_number, row in enumerate(rows, start=1):
            if row_number == header_row:
                header = row
                break

        if header is None:
            raise ValueError(f"工作表不足 {header_row} 行，无法找到表头")

        header_map = {
            cell_text(column_name): index
            for index, column_name in enumerate(header)
            if cell_text(column_name)
        }
        missing = [name for name in REQUIRED_COLUMNS if name not in header_map]
        if missing:
            raise ValueError(f"馆藏表缺少必要字段：{', '.join(missing)}")

        for row in rows:
            values: dict[str, str] = {}
            for column_name in REQUIRED_COLUMNS:
                index = header_map[column_name]
                values[column_name] = (
                    cell_text(row[index]) if index < len(row) else ""
                )
            yield values
    finally:
        workbook.close()


def merge_catalog(rows: Iterable[dict[str, str]]) -> tuple[list[CatalogRecord], int]:
    """按书目记录合并复本，返回书目列表和原始有效册数。

    优先使用“元数据记录号”作为聚合键。若某行记录号为空，则使用书名、
    责任者、出版社和出版年组合成后备键，避免把所有空编号记录错误合并。
    """

    records: "OrderedDict[str, CatalogRecord]" = OrderedDict()
    source_copy_count = 0

    for row in rows:
        title = row["题名"]
        if not title:
            # 没有题名的行无法用于 OCR 纠错，跳过但不让程序中断。
            continue

        metadata_id = row["元数据记录号"]
        fallback_key = "|".join(
            (title, row["责任者"], row["出版社"], row["出版年"])
        )
        key = f"id:{metadata_id}" if metadata_id else f"fallback:{fallback_key}"

        source_copy_count += 1
        is_available = row["书刊状态"] == "在架"

        if key not in records:
            records[key] = CatalogRecord(
                metadata_id=metadata_id,
                title=title,
                title_normalized=normalize_title(title),
                author=row["责任者"],
                call_number=row["索书号"],
                publisher=row["出版社"],
                publication_year=row["出版年"],
                standard_number=row["标准号"],
                document_type=row["文献类型"],
                holding_location=row["馆藏地"],
            )

        record = records[key]
        record.total_copy_count += 1
        record.available_copy_count += int(is_available)

        # 同一书目的不同册可能有部分字段缺失，合并时补齐首个非空值。
        record.author = choose_first(record.author, row["责任者"])
        record.call_number = choose_first(record.call_number, row["索书号"])
        record.publisher = choose_first(record.publisher, row["出版社"])
        record.publication_year = choose_first(
            record.publication_year, row["出版年"]
        )
        record.standard_number = choose_first(
            record.standard_number, row["标准号"]
        )

    return list(records.values()), source_copy_count


def write_catalog(records: list[CatalogRecord], destination: Path) -> None:
    """将精简馆藏目录写为 UTF-8 BOM CSV。"""

    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(record.to_row() for record in records)


def parse_args() -> argparse.Namespace:
    """解析命令行参数。"""

    parser = argparse.ArgumentParser(
        description="将按册馆藏 XLSX 转换为适合 OCR 匹配的精简 CSV。"
    )
    parser.add_argument(
        "--source",
        type=Path,
        default=PROJECT_ROOT
        / Path(
            "泰达西区库-馆藏清单-按册20260522-005/"
            "泰达西区库-馆藏清单-按册20260522-005.xlsx"
        ),
        help="原始馆藏 XLSX 路径。",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=PROJECT_ROOT / "data/processed/catalog/catalog.csv",
        help="精简 CSV 输出路径。",
    )
    parser.add_argument(
        "--sheet",
        default="Sheet0",
        help="需要读取的工作表名称。",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=5,
        help="表头所在行号，从 1 开始计数。",
    )
    return parser.parse_args()


def main() -> None:
    """执行馆藏目录转换并输出可核对的统计信息。"""

    args = parse_args()
    if not args.source.is_file():
        raise FileNotFoundError(f"找不到馆藏 Excel：{args.source.resolve()}")

    records, source_copy_count = merge_catalog(
        read_catalog_rows(args.source, args.sheet, args.header_row)
    )
    if not records:
        raise RuntimeError("未读取到任何有效题名，已停止生成空 CSV")

    write_catalog(records, args.output)
    print(f"原始有效馆藏册数：{source_copy_count}")
    print(f"合并后书目记录数：{len(records)}")
    print(f"输出文件：{args.output.resolve()}")


if __name__ == "__main__":
    main()
